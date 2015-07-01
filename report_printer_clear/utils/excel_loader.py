from copy import deepcopy
from decimal import Decimal
import openpyxl
import xlrd
from abc import ABCMeta, abstractmethod


class ExcelLoader():
    __metaclass__ = ABCMeta

    def __init__(self, filename):
        self.book = self.open_book(filename)

    def load(self, sheet_index, border):
        border = self.get_border(border)
        total_data = []
        for row_index in xrange(border['top'], border['bottom']+1):
            row = []
            for column_index in xrange(border['left'], border['right']+1):
                src_value = self.get_cell(sheet_index, row_index, column_index)
                if src_value:
                    value = Decimal(src_value)
                else:
                    value = Decimal(0)
                row.append(value)
            total_data.append(row)
        return total_data

    @abstractmethod
    def open_book(self, filename):
        pass

    @abstractmethod
    def get_border(self, border):
        pass

    @abstractmethod
    def get_cell(self, sheet_index, row_index, column_index):
        pass

    @abstractmethod
    def row_index(self, idx):
        pass

    @abstractmethod
    def column_index(self, idx):
        pass


class XlsxExcelLoader(ExcelLoader):

    def open_book(self, filename):
        return openpyxl.load_workbook(filename)

    def get_border(self, border):
        return border

    def get_cell(self, sheet_index, row_index, column_index):
        return self.book.worksheets[sheet_index].cell(row=row_index,
                                                      column=column_index).value

    def row_index(self, idx):
        return idx + 1

    def column_index(self, idx):
        return idx + 1


class XlsExcelLoader(ExcelLoader):

    def open_book(self, filename):
        return xlrd.open_workbook(filename)

    def get_border(self, border):
        new_border = deepcopy(border)
        new_border['top'] = border['top'] - 1
        new_border['bottom'] = border['bottom'] - 1
        new_border['left'] = border['left'] - 1
        new_border['right'] = border['right'] - 1
        return new_border

    def get_cell(self, sheet_index, row_index, column_index):
        return self.book.sheet_by_index(sheet_index).cell(row_index, column_index).value

    def row_index(self, idx):
        return idx

    def column_index(self, idx):
        return idx
