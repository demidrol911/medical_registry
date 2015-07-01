#! -*- coding: utf-8 -*-
from copy import deepcopy
from decimal import Decimal, InvalidOperation
import openpyxl

import os
import xlrd
import re

from abc import ABCMeta, abstractmethod


class ExcelLoader():
    __metaclass__ = ABCMeta

    def __init__(self):
        pass

    def load(self, filepath, sheet_index, border):
        print filepath
        self.open_book(filepath)
        border = self.get_border(border)
        border['bottom'] = self.length_row(sheet_index)
        sum_data = []
        for row_index in xrange(border['top'], border['bottom']+1):
            row = []
            for column_index in xrange(border['left'], border['right']+1):
                value_src = self.get_cell(sheet_index, row_index, column_index)
                try:
                    if value_src:
                        #print value_src
                        value = float(value_src)
                    else:
                        value = float(0)
                except:
                    value = value_src
                #print value
                row.append(value)
            sum_data.append(row)
        return sum_data

    def ignore_services(self, filepath):
        sheet_index = 1
        self.open_book(filepath)
        service_code = []
        for idx in xrange(self.length_row(sheet_index)):
            service_code.append(self.get_cell(sheet_index, self.row_index(idx), self.column_index(0)))
        return service_code

    @abstractmethod
    def open_book(self, filepath):
        pass

    @abstractmethod
    def get_border(self, border):
        pass

    @abstractmethod
    def get_cell(self, sheet_index, row_index, column_index):
        pass

    @abstractmethod
    def length_row(self, sheet_index):
        pass

    @abstractmethod
    def row_index(self, idx):
        pass

    @abstractmethod
    def column_index(self, idx):
        pass


class XlsxExcelLoader(ExcelLoader):

    def __init__(self):
        self.book = None

    def open_book(self, filepath):
        self.book = openpyxl.load_workbook(filepath)

    def get_border(self, border):
        return border

    def get_cell(self, sheet_index, row_index, column_index):
        return self.book.worksheets[sheet_index].cell(row=row_index, column=column_index).value

    def length_row(self, sheet_index):
        return len(self.book.worksheets[sheet_index].rows)

    def row_index(self, idx):
        return idx + 1

    def column_index(self, idx):
        return idx + 1


class XlsExcelLoader(ExcelLoader):

    def __init__(self):
        self.book = None

    def open_book(self, filepath):
        self.book = xlrd.open_workbook(filepath)

    def get_border(self, border):
        new_border = deepcopy(border)
        new_border['top'] = border['top'] - 1
        new_border['bottom'] = border['bottom'] - 1
        new_border['left'] = border['left'] - 1
        new_border['right'] = border['right'] - 1
        return new_border

    def get_cell(self, sheet_index, row_index, column_index):
        return self.book.sheet_by_index(sheet_index).cell(row_index, column_index).value

    def length_row(self, sheet_index):
        return self.book.sheet_by_index(sheet_index).nrows

    def row_index(self, idx):
        return idx

    def column_index(self, idx):
        return idx


def print_data(data):
    #print len(data), len(data[0])
    for row in data:
        for value in row:
            print value, #.replace('.', ','),
        print


def get_equal_filename(path_to_dir, filename):
    filename_pattern_parse = re.compile(ur'^(?P<mo>[\D1234]+?)(?:_?)(?P<sequence_number>\d?)\.xlsx$')
    parse_match = filename_pattern_parse.match(filename)

    if parse_match:
        filename_unique = parse_match.group('mo')
    else:
        filename_unique = ''

    filename_pattern = re.compile(ur'^%s(?:_?)(?P<sequence_number>\d*?)\.xlsx$' %
                                  (filename_unique, ))

    equals_filenames = []
    for dst_filename in os.listdir(path_to_dir):
        if filename_pattern.match(dst_filename):
            equals_filenames.append(dst_filename)
    return equals_filenames


def diff(data_src, data_dst):
    is_equals = True
    for i, row_src in enumerate(data_src):
        for j, value_src in enumerate(row_src):
            if abs(value_src - data_dst[i][j]) > 0.05:
                is_equals = False
    return is_equals


def diff_1(data_src, data_dst):
    is_equals = True
    for i, row_src in enumerate(data_src):
        if abs(row_src - data_dst[i]) > 0.05:
            is_equals = False
    return is_equals


def diff_dict(data_src, data_dst):
    is_equals = True
    for src_key in data_src:
        if src_key in data_dst:
            #print data_src[src_key],  data_dst[src_key]
            is_equals = diff_1(data_src[src_key], data_dst[src_key])
        else:
            is_equals = False
    return is_equals


def array_to_dict(data):
    result = {}
    for row in data:
        result[row[0]] = row[2:]
    return result


class ReportLoaderParameters():

    def __init__(self):
        self.sheet_index = 0
        self.top = 0
        self.left = 0
        self.right = 0
        self.bottom = 0

if __name__ == '__main__':
    DIR_SRC = ur'C:\TEST\OLD\G2015\Period04'
    DIR_DST = ur'C:\TEST\NEW\G2015\Period04'
    border = {
        'top': 7,
        'bottom': 7,
        'left': 1,
        'right': 27
    }

    for src_filename in os.listdir(DIR_SRC):
        #print src_filename, src_filename[-4:]
        equals_filenames = get_equal_filename(DIR_DST, src_filename)
        if equals_filenames:
            #print equals_filenames
            if src_filename[-4:] == 'xlsx':
                excel_loader = XlsxExcelLoader()
            elif src_filename[-3:] == 'xls':
                excel_loader = XlsExcelLoader()
            #border['bottom'] = excel_loader.length_row(2)
            data_src = excel_loader.load(os.path.join(DIR_SRC, src_filename), 2, border)
            #print_data(data_src)
            #print data_src

            #print_data(data_src)
        for dst_filename in equals_filenames:
            if dst_filename[-4:] == 'xlsx':
                excel_loader = XlsxExcelLoader()
            elif dst_filename[-3:] == 'xls':
                excel_loader = XlsExcelLoader()
            #border['bottom'] = excel_loader.length_row(2)
            #print os.path.join(DIR_DST, dst_filename)
            #print os.path.join(DIR_SRC, src_filename)
            data_dst = excel_loader.load(os.path.join(DIR_DST, dst_filename), 2, border)
            #print '-->', dst_filename
            #print_data(data_dst)

            is_equals = diff_dict(array_to_dict(data_src), array_to_dict(data_dst))

            if is_equals:
                pass
                '''
                print src_filename, dst_filename, 'YES'
                print_data(data_src)
                print_data(data_dst)
                print '-'*70
                '''

                #pass
            else:
                print src_filename, dst_filename, 'NO'
                print_data(data_src)
                print_data(data_dst)
                print '-'*70
                pass
                #print src_filename, dst_filename, 'No'
                #break
